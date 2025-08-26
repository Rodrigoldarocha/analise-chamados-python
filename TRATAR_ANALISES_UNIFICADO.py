# -*- coding: utf-8 -*-
"""
Pipeline STD - Tratamento, Métricas de SLA (equivalentes DAX), Análises e Exportação Excel
Versão Completa com todas as medidas DAX implementadas
"""

import logging
import warnings
import os
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
from datetime import datetime, timedelta
import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import calendar
import re

warnings.filterwarnings("ignore")

# ======================
# LOGGING
# ======================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)

# ======================
# CONFIGURAÇÕES
# ======================
class Config:
    """Configurações da aplicação"""

    # Mapeamento completo de Divisões (todos os estados)
    DIVISOES = {
        "DIV 01": {"GO": ["GO 01"], "UFs": ["AL", "CE", "PB", "PE", "RN"]},
        "DIV 02": {"GO": ["GO 01", "GO 02"], "UFs": ["BA", "SE"]},
        "DIV 03": {"GO": ["GO 01"], "UFs": ["CE", "MA", "PI"]},
        "DIV 04": {"GO": ["GO 02"], "UFs": ["AP", "PA"]},
        "DIV 05": {"GO": ["GO 02"], "UFs": ["AM", "RO", "RR", "AC"]},
        "DIV 06": {"GO": ["GO 02"], "UFs": ["DF", "GO"]},
        "DIV 07": {"GO": ["GO 02"], "UFs": ["MT", "MS"]},
        "DIV 08": {"GO": ["GO 03"], "UFs": ["SP"]},
        "DIV 09": {"GO": ["GO 03"], "UFs": ["RJ", "MG"]},
        "DIV 10": {"GO": ["GO 03"], "UFs": ["ES", "PR", "SC", "RS"]},
    }

    # Colunas importantes
    COLUNAS_IMPORTANTES = [
        "Divisão", "Gerência Operacional", "UF", "Base", "Origem", "Classificacao",
        "Solicitante", "Data_Criacao", "Data_Chegada", "Data_Previsao_Conclusao",
        "Data_Previsao_Chegada", "Data_Conclusao", "Data_de_Fechamento", "Fila",
        "Grupo", "Numero_Chamado", "Prioridade", "Substatus", "Status", "SubTipo",
        "Tipo", "Valor_Total", "Negocio", "Local_Nome", "Uniorg_Comercial",
        "Tempo_de_Custo", "Data_do_Primeiro_Encaminhamento", "Fornecedor",
        "Nota_Inicial", "originador", "regional", "Responsavel", "prazo_inicio",
        "prazo_conclusao", "rede", "modulo", "DURAÇÃO CHAMADO", "Duração Chamado"
    ]

    # Colunas de data
    DATE_COLUMNS = [
        "Data_Criacao", "Data_Chegada", "Data_Previsao_Conclusao",
        "Data_Previsao_Chegada", "Data_Conclusao", "Data_de_Fechamento",
        "Data_do_Primeiro_Encaminhamento"
    ]

    # Metas
    META_SLA = 96.0  # 96%
    META_LIMPEZA = 98.0  # 98%

    # Cores Excel
    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ORANGE_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

    # Caminhos (agora relativos ao diretório do script)
    SCRIPT_DIR = Path(__file__).parent
    FILE_PATH = SCRIPT_DIR / "Base Geral STD.xlsx"
    OUTPUT_BASE_TRATADA = SCRIPT_DIR / "Base_Tratada.xlsx"
    OUTPUT_ANALISE_COMPLETA = SCRIPT_DIR / "Analise_Chamados_Completa.xlsx"

    # Configurações de análise
    TOP_RESPONSABLES = 15
    DIAS_FECHAMENTO_PENDENTE = 30

# ======================
# UTILITÁRIOS
# ======================
class DateUtils:
    """Utilitários para manipulação de datas"""
    
    @staticmethod
    def is_business_day(date: datetime) -> bool:
        """Verifica se é dia útil"""
        if pd.isna(date):
            return False
        return date.weekday() < 5  # Segunda a sexta
    
    @staticmethod
    def business_days_between(start_date: datetime, end_date: datetime) -> int:
        """Calcula dias úteis entre duas datas"""
        if pd.isna(start_date) or pd.isna(end_date):
            return 0
        
        # Se as datas forem iguais, retorna 0
        if start_date.date() == end_date.date():
            return 0
            
        days = np.busday_count(start_date.date(), end_date.date())
        return max(0, days)
    
    @staticmethod
    def get_month_name(period) -> str:
        """Retorna nome do mês a partir de período"""
        if hasattr(period, 'month'):
            return calendar.month_name[period.month]
        return str(period)
    
    @staticmethod
    def format_time_duration(seconds: float) -> str:
        """Formata duração de tempo em formato HH:MM:SS"""
        if pd.isna(seconds):
            return "00:00:00"
        
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        secs = int(seconds % 60)
        
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"

# ======================
# PROCESSAMENTO
# ======================
class STDDataProcessor:
    """Carrega, prepara e computa colunas auxiliares e de SLA"""

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.df_original: pd.DataFrame = pd.DataFrame()
        self.df_processed: pd.DataFrame = pd.DataFrame()
        self._setup_mappings()
        self.stats: Dict[str, float] = {}
        self.calendario: pd.DataFrame = pd.DataFrame()

    def _setup_mappings(self) -> None:
        """Configura mapeamentos UF -> Divisão e GO"""
        self.uf_para_divisao = {}
        self.uf_para_go = {}
        
        for divisao, info in Config.DIVISOES.items():
            for uf in info["UFs"]:
                self.uf_para_divisao[uf] = divisao
                # Usa a primeira GO como padrão
                self.uf_para_go[uf] = info["GO"][0] if info["GO"] else "GO Não Definido"

    def load_data(self) -> pd.DataFrame:
        """Carrega dados do arquivo Excel"""
        logger.info(f"Lendo base: {self.file_path}")
        if not self.file_path.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {self.file_path}")
        
        try:
            # Tenta ler todas as sheets para encontrar a correta
            excel_file = pd.ExcelFile(self.file_path)
            sheet_name = excel_file.sheet_names[0]  # Pega a primeira sheet
            
            self.df_original = pd.read_excel(self.file_path, sheet_name=sheet_name)
            logger.info(f"Registros lidos: {len(self.df_original)}")
            logger.info(f"Colunas disponíveis: {list(self.df_original.columns)}")
            return self.df_original
        except Exception as e:
            logger.error(f"Erro ao ler arquivo: {e}")
            raise

    def prepare_data(self) -> pd.DataFrame:
        """Prepara e processa os dados"""
        if self.df_original.empty:
            raise ValueError("Dados não carregados. Use load_data().")

        logger.info("Preparando dados...")

        # Garantir colunas importantes
        self._ensure_columns()
        
        # Filtrar apenas colunas importantes
        available_cols = [col for col in Config.COLUNAS_IMPORTANTES if col in self.df_original.columns]
        self.df_processed = self.df_original[available_cols].copy()
        
        # Aplicar mapeamentos
        self._apply_mappings()
        
        # Converter datas
        self._convert_dates()
        
        # Criar calendário
        self._create_calendar()
        
        # Criar colunas auxiliares (equivalentes DAX)
        self._create_dax_equivalent_columns()
        
        # Identificar chamados atrasados e em aberto
        self._identify_late_and_open_calls()
        
        # Cálculos de SLA
        self._calculate_sla_status()
        
        # Calcular estatísticas
        self._compute_statistics()
        
        logger.info("Dados preparados com sucesso.")
        return self.df_processed

    def _ensure_columns(self) -> None:
        """Garante que todas as colunas importantes existam no DataFrame"""
        for col in Config.COLUNAS_IMPORTANTES:
            if col not in self.df_original.columns:
                logger.warning(f"Coluna {col} não encontrada - criando vazia")
                self.df_original[col] = None

    def _apply_mappings(self) -> None:
        """Aplica mapeamentos UF -> Divisão e GO"""
        if "UF" in self.df_processed.columns:
            self.df_processed["Divisão"] = self.df_processed["UF"].map(
                self.uf_para_divisao
            ).fillna("Divisão Não Definida")
            
            self.df_processed["Gerência Operacional"] = self.df_processed["UF"].map(
                self.uf_para_go
            ).fillna("GO Não Definida")
        else:
            logger.warning("Coluna UF não encontrada - não foi possível mapear Divisão e GO")

    def _convert_dates(self) -> None:
        """Converte colunas de data para datetime"""
        for col in Config.DATE_COLUMNS:
            if col in self.df_processed.columns:
                self.df_processed[col] = pd.to_datetime(
                    self.df_processed[col], errors='coerce', dayfirst=True
                )
                # Log de conversão
                n_converted = self.df_processed[col].notna().sum()
                logger.info(f"Coluna {col}: {n_converted} datas convertidas")

    def _create_calendar(self) -> None:
        """Cria calendário baseado nas datas dos chamados"""
        if "Data_Criacao" not in self.df_processed.columns:
            logger.warning("Coluna Data_Criacao não encontrada - não foi possível criar calendário")
            return
            
        dt_min = self.df_processed["Data_Criacao"].min()
        dt_max = self.df_processed["Data_Criacao"].max()
        
        if pd.isna(dt_min) or pd.isna(dt_max):
            logger.warning("Datas inválidas - não foi possível criar calendário")
            return
            
        # Criar calendário com todas as datas entre dt_min e dt_max
        self.calendario = pd.DataFrame({
            "Date": pd.date_range(start=dt_min, end=dt_max)
        })
        
        # Adicionar colunas auxiliares ao calendário
        self.calendario["Ano"] = self.calendario["Date"].dt.year
        self.calendario["Mes"] = self.calendario["Date"].dt.month
        self.calendario["Nome_Mes"] = self.calendario["Date"].dt.month_name()
        self.calendario["Dia_Semana"] = self.calendario["Date"].dt.day_name()
        self.calendario["Dia_Util"] = np.where(
            self.calendario["Dia_Semana"].isin(["Saturday", "Sunday"]), "N", "S"
        )
        self.calendario["Semana_Ano"] = self.calendario["Date"].dt.isocalendar().week
        self.calendario["Mes_Ano"] = self.calendario["Date"].dt.to_period("M").astype(str)

    def _create_dax_equivalent_columns(self) -> None:
        """Cria colunas equivalentes às medidas DAX do Power BI"""
        df = self.df_processed
        hoje = pd.to_datetime("today").normalize()
        
        # 1. Prazo Ajustado (equivalente às medidas DAX)
        if "prazo_inicio" in df.columns:
            df["Prazo_Inicio_Ajustado"] = np.where(
                df["prazo_inicio"].astype(str).str.upper() == "NA", 
                "NP", 
                df["prazo_inicio"]
            )
        else:
            df["Prazo_Inicio_Ajustado"] = "NP"  # Valor padrão se a coluna não existir
            logger.warning("Coluna prazo_inicio não encontrada - usando valor padrão 'NP'")
        
        if "prazo_conclusao" in df.columns:
            df["Prazo_Conclusao_Ajustado"] = np.where(
                df["prazo_conclusao"].astype(str).str.upper() == "NA", 
                "NP", 
                df["prazo_conclusao"]
            )
        else:
            df["Prazo_Conclusao_Ajustado"] = "NP"  # Valor padrão se a coluna não existir
            logger.warning("Coluna prazo_conclusao não encontrada - usando valor padrão 'NP'")
        
        # 2. Status Chamado (equivalente DAX)
        df["Status_Chamado"] = np.where(
            df["Data_Conclusao"].isna(), 
            "Pendente", 
            "Concluído"
        )
        
        # 3. Status Fechamento (equivalente DAX)
        df["Status_Fechamento"] = np.where(
            df["Data_Conclusao"].isna() & df["Data_de_Fechamento"].isna(),
            "Pendente", 
            "Concluído"
        )
        
        # 4. Status Financeiro (equivalente DAX)
        df["Status_Financeiro"] = np.where(
            df["Data_de_Fechamento"].isna(),
            "Pendente",
            "Fechado"
        )
        
        # 5. Estoque Atual (equivalente DAX)
        df["Estoque_Atual"] = np.where(
            df["Data_de_Fechamento"].isna() & df["Data_Conclusao"].isna(),
            "Estoque Atual", 
            "Não"
        )
        
        # 6. Data Conclusão Ajustada (equivalente DAX)
        df["Data_Conclusao_Ajustada"] = np.where(
            df["Data_de_Fechamento"].notna() & df["Data_Conclusao"].isna(),
            df["Data_de_Fechamento"],
            df["Data_Conclusao"]
        )
        
        # 7. Data Estoque (equivalente DAX)
        df["Data_Estoque"] = np.where(
            df["Estoque_Atual"] == "Estoque Atual",
            hoje,
            pd.NaT
        )
        
        # 8. Fechamento Pendente (equivalente DAX)
        df["Fechamento_Pendente"] = np.where(
            df["Data_de_Fechamento"].isna() & 
            df["Data_Conclusao"].notna() &
            ((hoje - df["Data_Conclusao"]).dt.days > Config.DIAS_FECHAMENTO_PENDENTE),
            "Sim", 
            "Não"
        )
        
        # 9. Duração do Chamado (equivalente DAX)
        df["DURACAO_CHAMADO"] = (hoje - df["Data_Criacao"]).dt.days
        
        # 10. Dias de Atraso (equivalente DAX)
        df["Dias_Atrasos"] = np.where(
            df["Data_Previsao_Conclusao"].notna(),
            (hoje - df["Data_Previsao_Conclusao"]).dt.days + 1,
            0
        )
        
        # 11. Dias Chegada (equivalente DAX)
        df["Dias_Chegada"] = np.where(
            df["Data_Chegada"].notna(),
            (df["Data_Chegada"] - df["Data_Criacao"]).dt.days,
            0
        )
        df["Dias_Chegada"] = np.where(df["Dias_Chegada"] == 0, 1, df["Dias_Chegada"])
        
        # 12. Dias Conclusão (equivalente DAX)
        df["Dias_Conclusao"] = np.where(
            df["Data_Conclusao"].notna(),
            (df["Data_Conclusao"] - df["Data_Criacao"]).dt.days,
            0
        )
        df["Dias_Conclusao"] = np.where(df["Dias_Conclusao"] == 0, 1, df["Dias_Conclusao"])
        
        # 13. Dias Fechados (equivalente DAX)
        df["Dias_Fechados"] = np.where(
            df["Data_de_Fechamento"].notna(),
            (df["Data_de_Fechamento"] - df["Data_Criacao"]).dt.days,
            0
        )
        df["Dias_Fechados"] = np.where(df["Dias_Fechados"] == 0, 1, df["Dias_Fechados"])
        
        # 14. Tempo Atendimento (equivalente DAX)
        df["Tempo_Atendimento"] = np.where(
            df["Data_Conclusao"].notna(),
            (df["Data_Conclusao"] - df["Data_Criacao"]).dt.days,
            0
        )
        
        # 15. Horas Chegada x Criação (para cálculo de tempo médio)
        df["Horas_Chegada_x_Criacao"] = np.where(
            df["Data_Chegada"].notna(),
            (df["Data_Chegada"] - df["Data_Criacao"]).dt.total_seconds(),
            0
        )
        
        # 16. Faixa Dias em Aberto (equivalente DAX)
        conditions = [
            (df["DURACAO_CHAMADO"] > 90) & df["Data_Conclusao"].notna(),
            (df["DURACAO_CHAMADO"] > 60) & df["Data_Conclusao"].notna(),
            (df["DURACAO_CHAMADO"] > 30) & df["Data_Conclusao"].notna()
        ]
        choices = ["+90 dias", "+60 dias", "+30 dias"]
        df["Faixa_Dias_em_Aberto"] = np.select(conditions, choices, default="-30 dias")
        
        # 17. À VENCER WTM 30 DIAS (equivalente DAX)
        df["A_VENCER_WTM_30_DIAS"] = np.where(
            (df["DURACAO_CHAMADO"] < 30) & (df["DURACAO_CHAMADO"] >= 20),
            "À VENCER WTM +30 DIAS", 
            "OUTROS"
        )
        
        # 18. UF - Mapa (equivalente DAX)
        df["UF_Mapa"] = df["UF"].astype(str) + "-" + "Brasil"
        
        # Período (Mês/Ano) para análises temporais
        if "Data_Criacao" in df.columns:
            df["Mes_Ano"] = df["Data_Criacao"].dt.to_period("M")
            df["Ano"] = df["Data_Criacao"].dt.year
            df["Mes"] = df["Data_Criacao"].dt.month
            df["Nome_Mes"] = df["Mes_Ano"].apply(DateUtils.get_month_name)

    def _identify_late_and_open_calls(self) -> None:
        """Identifica chamados atrasados e em aberto"""
        df = self.df_processed
        hoje = pd.to_datetime("today").normalize()
        
        # Status de Atraso
        conditions = [
            # Chamados concluídos mas com atraso
            (df["Data_Conclusao"].notna() & df["Data_Previsao_Conclusao"].notna() & 
             (df["Data_Conclusao"] > df["Data_Previsao_Conclusao"])),
            
            # Chamados não concluídos e com previsão vencida
            (df["Data_Conclusao"].isna() & df["Data_Previsao_Conclusao"].notna() & 
             (df["Data_Previsao_Conclusao"] < hoje)),
            
            # Chamados não concluídos e sem previsão
            (df["Data_Conclusao"].isna() & df["Data_Previsao_Conclusao"].isna()),
            
            # Chamados em dia (não atrasados)
            ((df["Data_Conclusao"].notna() & df["Data_Previsao_Conclusao"].notna() & 
              (df["Data_Conclusao"] <= df["Data_Previsao_Conclusao"])) |
             (df["Data_Conclusao"].isna() & df["Data_Previsao_Conclusao"].notna() & 
              (df["Data_Previsao_Conclusao"] >= hoje)))
        ]
        
        choices = [
            "Concluído com Atraso",
            "Atrasado",
            "Em Aberto (Sem Previsão)",
            "Em Dia"
        ]
        
        df["Status_Atraso"] = np.select(conditions, choices, default="Status Indefinido")
        
        # Dias em Atraso
        df["Dias_Atraso"] = 0
        
        # Para chamados concluídos com atraso
        mask_concluido_atraso = (
            df["Data_Conclusao"].notna() & 
            df["Data_Previsao_Conclusao"].notna() & 
            (df["Data_Conclusao"] > df["Data_Previsao_Conclusao"])
        )
        df.loc[mask_concluido_atraso, "Dias_Atraso"] = (
            df["Data_Conclusao"] - df["Data_Previsao_Conclusao"]
        ).dt.days
        
        # Para chamados não concluídos e com previsão vencida
        mask_nao_concluido_atraso = (
            df["Data_Conclusao"].isna() & 
            df["Data_Previsao_Conclusao"].notna() & 
            (df["Data_Previsao_Conclusao"] < hoje)
        )
        df.loc[mask_nao_concluido_atraso, "Dias_Atraso"] = (
            hoje - df["Data_Previsao_Conclusao"]
        ).dt.days
        
        # Para chamados não concluídos e sem previsão
        mask_sem_previsao = (
            df["Data_Conclusao"].isna() & 
            df["Data_Previsao_Conclusao"].isna()
        )
        df.loc[mask_sem_previsao, "Dias_Atraso"] = (
            hoje - df["Data_Criacao"]
        ).dt.days

    def _calculate_sla_status(self) -> None:
        """Calcula status de SLA para início e conclusão"""
        df = self.df_processed
        
        # SLA Início
        def _calc_sla_inicio(row):
            if pd.isna(row.get("Data_Previsao_Chegada")):
                return "Não Definido"
            
            data_inicio = row.get("Data_do_Primeiro_Encaminhamento")
            if pd.isna(data_inicio):
                data_inicio = row.get("Data_Chegada")
            if pd.isna(data_inicio):
                return "Não Definido"
                
            return "NP" if data_inicio <= row["Data_Previsao_Chegada"] else "FP"

        # SLA Conclusão
        def _calc_sla_conclusao(row):
            if pd.isna(row.get("Data_Previsao_Conclusao")):
                return "Não Definido"
            if pd.isna(row.get("Data_Conclusao")):
                return "Pendente"
                
            return "NP" if row["Data_Conclusao"] <= row["Data_Previsao_Conclusao"] else "FP"

        # Aplica as funções apenas se as colunas necessárias existem
        if all(col in df.columns for col in ["Data_Previsao_Chegada", "Data_do_Primeiro_Encaminhamento", "Data_Chegada"]):
            df["Status_Prazo_Inicio"] = df.apply(_calc_sla_inicio, axis=1)
        else:
            df["Status_Prazo_Inicio"] = "Não Definido"
            logger.warning("Colunas necessárias para Status_Prazo_Inicio não encontradas")
        
        if all(col in df.columns for col in ["Data_Previsao_Conclusao", "Data_Conclusao"]):
            df["Status_Prazo_Conclusao"] = df.apply(_calc_sla_conclusao, axis=1)
        else:
            df["Status_Prazo_Conclusao"] = "Não Definido"
            logger.warning("Colunas necessárias para Status_Prazo_Conclusao não encontradas")

        # Calcular dias de atraso
        if all(col in df.columns for col in ["Data_Chegada", "Data_Previsao_Chegada"]):
            df["Dias_Atraso_Inicio"] = np.where(
                df["Status_Prazo_Inicio"] == "FP",
                (df["Data_Chegada"] - df["Data_Previsao_Chegada"]).dt.days,
                0
            )
        
        if all(col in df.columns for col in ["Data_Conclusao", "Data_Previsao_Conclusao"]):
            df["Dias_Atraso_Conclusao"] = np.where(
                df["Status_Prazo_Conclusao"] == "FP",
                (df["Data_Conclusao"] - df["Data_Previsao_Conclusao"]).dt.days,
                0
            )

        # Duração em dias úteis
        if "Data_Criacao" in df.columns:
            df["Duracao_Chamado_Dias_Uteis"] = df.apply(
                lambda x: DateUtils.business_days_between(
                    x["Data_Criacao"], 
                    x["Data_Conclusao"] if pd.notna(x.get("Data_Conclusao")) else pd.Timestamp.now()
                ), axis=1
            )

    def _compute_statistics(self) -> None:
        """Calcula estatísticas equivalentes às medidas DAX"""
        df = self.df_processed
        
        # Total Chamados (equivalente DAX)
        total_chamados = len(df)
        
        # Total chamados termino (equivalente DAX)
        total_chamados_termino = df["Data_Conclusao"].notna().sum() if "Data_Conclusao" in df.columns else 0
        
        # Total Conclusão NP (equivalente DAX)
        total_conclusao_np = (df["Prazo_Conclusao_Ajustado"] == "NP").sum() if "Prazo_Conclusao_Ajustado" in df.columns else 0
        
        # Total Inicio NP (equivalente DAX)
        total_inicio_np = (df["Prazo_Inicio_Ajustado"] == "NP").sum() if "Prazo_Inicio_Ajustado" in df.columns else 0
        
        # SLA Início (equivalente DAX)
        sla_inicio = (total_inicio_np / total_chamados * 100) if total_chamados > 0 else 0
        
        # SLA Término (equivalente DAX)
        sla_termino = (total_conclusao_np / total_chamados_termino * 100) if total_chamados_termino > 0 else 0
        
        # Comparações com meta (equivalentes DAX)
        comparacao_meta_inicio = sla_inicio - Config.META_SLA
        comparacao_meta_termino = sla_termino - Config.META_SLA
        comparacao_meta_limpeza_termino = sla_termino - Config.META_LIMPEZA
        
        # Total Estoque (equivalente DAX)
        total_estoque = (df["Estoque_Atual"] == "Estoque Atual").sum() if "Estoque_Atual" in df.columns else 0
        
        # Total Fornecedor (equivalente DAX)
        total_fornecedor = df["Fornecedor"].nunique() if "Fornecedor" in df.columns else 0
        
        # Total Chamados Concluídos (equivalente DAX)
        total_chamados_concluidos = total_chamados_termino
        
        # Total chamados FP (equivalente DAX)
        total_chamados_fp = total_chamados - total_conclusao_np
        
        # Fechamento Pendente
        fechamento_pendente = (df["Fechamento_Pendente"] == "Sim").sum() if "Fechamento_Pendente" in df.columns else 0
        
        # À VENCER WTM 30 DIAS
        a_vencer_wtm = (df["A_VENCER_WTM_30_DIAS"] == "À VENCER WTM +30 DIAS").sum() if "A_VENCER_WTM_30_DIAS" in df.columns else 0
        
        # Chamados Atrasados e Em Aberto
        chamados_atrasados = (df["Status_Atraso"].isin(["Atrasado", "Concluído com Atraso"])).sum() if "Status_Atraso" in df.columns else 0
        chamados_em_aberto = (df["Status_Atraso"] == "Em Aberto (Sem Previsão)").sum() if "Status_Atraso" in df.columns else 0
        
        # Médias (equivalentes DAX)
        media_dias_atrasos = df["Dias_Atrasos"].mean() if "Dias_Atrasos" in df.columns else 0
        media_dias_chegada = df["Dias_Chegada"].mean() if "Dias_Chegada" in df.columns else 0
        media_dias_conclusao = df["Dias_Conclusao"].mean() if "Dias_Conclusao" in df.columns else 0
        media_dias_fechamento = df["Dias_Fechados"].mean() if "Dias_Fechados" in df.columns else 0
        media_tempo_atendimento = df["Tempo_Atendimento"].mean() if "Tempo_Atendimento" in df.columns else 0
        
        # Média Valor OS
        media_valor_os = df["Valor_Total"].mean() if "Valor_Total" in df.columns else 0
        
        # Quantidade de Agências
        qtd_agencias = df["Uniorg_Comercial"].nunique() if "Uniorg_Comercial" in df.columns else 0
        
        # Total Valor OS
        total_valor_os = df["Valor_Total"].sum() if "Valor_Total" in df.columns else 0
        
        # Média Tempo Chegada (formato HHMMSS)
        media_tempo_chegada_segundos = df["Horas_Chegada_x_Criacao"].mean() if "Horas_Chegada_x_Criacao" in df.columns else 0
        media_tempo_chegada = DateUtils.format_time_duration(media_tempo_chegada_segundos)
        
        # Tempo Chegada Total (formato HHMMSS)
        tempo_chegada_total_segundos = df["Horas_Chegada_x_Criacao"].sum() if "Horas_Chegada_x_Criacao" in df.columns else 0
        tempo_chegada_total = DateUtils.format_time_duration(tempo_chegada_total_segundos)
        
        self.stats = {
            "Total Chamados": total_chamados,
            "Total Chamados Termino": total_chamados_termino,
            "Total Conclusão NP": total_conclusao_np,
            "Total Inicio NP": total_inicio_np,
            "SLA Início": sla_inicio,
            "SLA Término": sla_termino,
            "Comparação Meta Inicio": comparacao_meta_inicio,
            "Comparação Meta Término": comparacao_meta_termino,
            "Comparação Meta Limpeza Término": comparacao_meta_limpeza_termino,
            "Total Estoque": total_estoque,
            "Total Fornecedor": total_fornecedor,
            "Total Chamados Concluídos": total_chamados_concluidos,
            "Total Chamados FP": total_chamados_fp,
            "Fechamento Pendente": fechamento_pendente,
            "À VENCER WTM 30 DIAS": a_vencer_wtm,
            "Chamados Atrasados": chamados_atrasados,
            "Chamados Em Aberto": chamados_em_aberto,
            "Media Dias Atrasos": media_dias_atrasos,
            "Média Dias Chegada": media_dias_chegada,
            "Média Dias Conclusão": media_dias_conclusao,
            "Média Dias Fechamento": media_dias_fechamento,
            "Media Tempo Atendimento": media_tempo_atendimento,
            "Média Valor OS": media_valor_os,
            "Qtd Agencias": qtd_agencias,
            "Total Valor OS": total_valor_os,
            "Media Tempo Chegada": media_tempo_chegada,
            "Tempo Chegada": tempo_chegada_total
        }

    def save_processed_data(self, output_path: str) -> None:
        """Salva dados processados"""
        if self.df_processed.empty:
            raise ValueError("Não há dados processados para salvar.")
        
        self.df_processed.to_excel(output_path, index=False, engine="openpyxl")
        logger.info(f"Base tratada salva em: {output_path}")

# ======================
# ANÁLISE
# ======================
class STDAnalyzer:
    """Cria tabelas e métricas de análise"""

    def __init__(self, df: pd.DataFrame):
        self.df = df
        self.results: Dict[str, pd.DataFrame] = {}

    def calculate_general_stats(self) -> Dict[str, float]:
        """Calcula estatísticas gerais equivalentes às medidas DAX"""
        df = self.df
        
        # Verifica se as colunas necessárias existem
        has_prazo_conclusao = "Prazo_Conclusao_Ajustado" in df.columns
        has_prazo_inicio = "Prazo_Inicio_Ajustado" in df.columns
        has_data_conclusao = "Data_Conclusao" in df.columns
        has_estoque = "Estoque_Atual" in df.columns
        has_fornecedor = "Fornecedor" in df.columns
        has_fechamento_pendente = "Fechamento_Pendente" in df.columns
        has_wtm = "A_VENCER_WTM_30_DIAS" in df.columns
        has_status_atraso = "Status_Atraso" in df.columns
        has_dias_atrasos = "Dias_Atrasos" in df.columns
        has_dias_chegada = "Dias_Chegada" in df.columns
        has_dias_conclusao = "Dias_Conclusao" in df.columns
        has_dias_fechados = "Dias_Fechados" in df.columns
        has_tempo_atendimento = "Tempo_Atendimento" in df.columns
        has_valor_total = "Valor_Total" in df.columns
        has_uniorg_comercial = "Uniorg_Comercial" in df.columns
        has_horas_chegada = "Horas_Chegada_x_Criacao" in df.columns
        
        total_chamados = len(df)
        total_chamados_termino = df["Data_Conclusao"].notna().sum() if has_data_conclusao else 0
        total_conclusao_np = (df["Prazo_Conclusao_Ajustado"] == "NP").sum() if has_prazo_conclusao else 0
        total_inicio_np = (df["Prazo_Inicio_Ajustado"] == "NP").sum() if has_prazo_inicio else 0
        
        sla_inicio = (total_inicio_np / total_chamados * 100) if total_chamados > 0 else 0
        sla_termino = (total_conclusao_np / total_chamados_termino * 100) if total_chamados_termino > 0 else 0
        
        # Chamados Atrasados e Em Aberto
        chamados_atrasados = (df["Status_Atraso"].isin(["Atrasado", "Concluído com Atraso"])).sum() if has_status_atraso else 0
        chamados_em_aberto = (df["Status_Atraso"] == "Em Aberto (Sem Previsão)").sum() if has_status_atraso else 0
        
        # Médias (equivalentes DAX)
        media_dias_atrasos = df["Dias_Atrasos"].mean() if has_dias_atrasos else 0
        media_dias_chegada = df["Dias_Chegada"].mean() if has_dias_chegada else 0
        media_dias_conclusao = df["Dias_Conclusao"].mean() if has_dias_conclusao else 0
        media_dias_fechamento = df["Dias_Fechados"].mean() if has_dias_fechados else 0
        media_tempo_atendimento = df["Tempo_Atendimento"].mean() if has_tempo_atendimento else 0
        
        # Média Valor OS
        media_valor_os = df["Valor_Total"].mean() if has_valor_total else 0
        
        # Quantidade de Agências
        qtd_agencias = df["Uniorg_Comercial"].nunique() if has_uniorg_comercial else 0
        
        # Total Valor OS
        total_valor_os = df["Valor_Total"].sum() if has_valor_total else 0
        
        # Média Tempo Chegada (formato HHMMSS)
        media_tempo_chegada_segundos = df["Horas_Chegada_x_Criacao"].mean() if has_horas_chegada else 0
        media_tempo_chegada = DateUtils.format_time_duration(media_tempo_chegada_segundos)
        
        # Tempo Chegada Total (formato HHMMSS)
        tempo_chegada_total_segundos = df["Horas_Chegada_x_Criacao"].sum() if has_horas_chegada else 0
        tempo_chegada_total = DateUtils.format_time_duration(tempo_chegada_total_segundos)
        
        stats = {
            "Total Chamados": total_chamados,
            "Total Chamados Termino": total_chamados_termino,
            "Total Conclusão NP": total_conclusao_np,
            "Total Inicio NP": total_inicio_np,
            "SLA Início": sla_inicio,
            "SLA Término": sla_termino,
            "Comparação Meta Inicio": sla_inicio - Config.META_SLA,
            "Comparação Meta Término": sla_termino - Config.META_SLA,
            "Comparação Meta Limpeza Término": sla_termino - Config.META_LIMPEZA,
            "Total Estoque": (df["Estoque_Atual"] == "Estoque Atual").sum() if has_estoque else 0,
            "Total Fornecedor": df["Fornecedor"].nunique() if has_fornecedor else 0,
            "Total Chamados Concluídos": total_chamados_termino,
            "Total Chamados FP": total_chamados - total_conclusao_np,
            "Fechamento Pendente": (df["Fechamento_Pendente"] == "Sim").sum() if has_fechamento_pendente else 0,
            "À VENCER WTM 30 DIAS": (df["A_VENCER_WTM_30_DIAS"] == "À VENCER WTM +30 DIAS").sum() if has_wtm else 0,
            "Chamados Atrasados": chamados_atrasados,
            "Chamados Em Aberto": chamados_em_aberto,
            "Media Dias Atrasos": media_dias_atrasos,
            "Média Dias Chegada": media_dias_chegada,
            "Média Dias Conclusão": media_dias_conclusao,
            "Média Dias Fechamento": media_dias_fechamento,
            "Media Tempo Atendimento": media_tempo_atendimento,
            "Média Valor OS": media_valor_os,
            "Qtd Agencias": qtd_agencias,
            "Total Valor OS": total_valor_os,
            "Media Tempo Chegada": media_tempo_chegada,
            "Tempo Chegada": tempo_chegada_total
        }
        
        self.results["estatisticas_gerais"] = pd.DataFrame.from_dict(
            stats, orient='index', columns=['Valor']
        )
        return stats

    def analyze_by_dimension(self, dimension: str, top_n: int = 20) -> pd.DataFrame:
        """Analisa por dimensão específica"""
        if dimension not in self.df.columns:
            logger.warning(f"Dimensão {dimension} não encontrada")
            return pd.DataFrame()
        
        # Verifica se as colunas necessárias existem
        has_numero_chamado = "Numero_Chamado" in self.df.columns
        has_prazo_inicio = "Prazo_Inicio_Ajustado" in self.df.columns
        has_prazo_conclusao = "Prazo_Conclusao_Ajustado" in self.df.columns
        has_duracao = "Duracao_Chamado_Dias_Uteis" in self.df.columns
        has_valor_total = "Valor_Total" in self.df.columns
        
        if not all([has_numero_chamado, has_prazo_inicio, has_prazo_conclusao]):
            logger.warning(f"Colunas necessárias não encontradas para análise por {dimension}")
            return pd.DataFrame()
        
        analysis = (
            self.df.groupby(dimension)
            .agg(
                Total_Chamados=('Numero_Chamado', 'count'),
                NP_Inicio=('Prazo_Inicio_Ajustado', lambda x: (x == 'NP').sum()),
                NP_Conclusao=('Prazo_Conclusao_Ajustado', lambda x: (x == 'NP').sum()),
                Tempo_Medio_Resolucao=('Duracao_Chamado_Dias_Uteis', 'mean') if has_duracao else ('Numero_Chamado', 'count'),
                Valor_Total_OS=('Valor_Total', 'sum') if has_valor_total else ('Numero_Chamado', 'count')
            )
            .round(2)
            .sort_values('Total_Chamados', ascending=False)
            .head(top_n)
        )
        
        analysis['% SLA Início'] = (analysis['NP_Inicio'] / analysis['Total_Chamados'] * 100).round(2)
        analysis['% SLA Conclusão'] = (analysis['NP_Conclusao'] / analysis['Total_Chamados'] * 100).round(2)
        
        return analysis

    def analyze_monthly_evolution(self) -> pd.DataFrame:
        """Analisa evolução mensal"""
        if not all(col in self.df.columns for col in ['Ano', 'Mes', 'Nome_Mes', 'Numero_Chamado', 'Prazo_Inicio_Ajustado', 'Prazo_Conclusao_Ajustado']):
            logger.warning("Colunas necessárias não encontradas para análise mensal")
            return pd.DataFrame()
        
        monthly = (
            self.df.groupby(['Ano', 'Mes', 'Nome_Mes'])
            .agg(
                Total_Chamados=('Numero_Chamado', 'count'),
                NP_Inicio=('Prazo_Inicio_Ajustado', lambda x: (x == 'NP').sum()),
                NP_Conclusao=('Prazo_Conclusao_Ajustado', lambda x: (x == 'NP').sum())
            )
            .reset_index()
        )
        
        monthly['% SLA Início'] = (monthly['NP_Inicio'] / monthly['Total_Chamados'] * 100).round(2)
        monthly['% SLA Conclusão'] = (monthly['NP_Conclusao'] / monthly['Total_Chamados'] * 100).round(2)
        monthly['Período'] = monthly['Nome_Mes'] + ' ' + monthly['Ano'].astype(str)
        
        return monthly[['Período', 'Total_Chamados', 'NP_Inicio', 'NP_Conclusao', '% SLA Início', '% SLA Conclusão']]

    def get_top_responsibles(self, top_n: int = Config.TOP_RESPONSABLES) -> pd.DataFrame:
        """Obtém top responsáveis"""
        if "Responsavel" not in self.df.columns:
            logger.warning("Coluna Responsavel não encontrada")
            return pd.DataFrame()
        
        top = (
            self.df['Responsavel']
            .value_counts()
            .head(top_n)
            .reset_index()
        )
        top.columns = ['Responsavel', 'Total_Chamados']
        return top

    def get_fp_analysis(self) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """Analisa chamados FP"""
        fp_inicio = self.df[self.df['Status_Prazo_Inicio'] == 'FP'].copy() if 'Status_Prazo_Inicio' in self.df.columns else pd.DataFrame()
        fp_conclusao = self.df[self.df['Status_Prazo_Conclusao'] == 'FP'].copy() if 'Status_Prazo_Conclusao' in self.df.columns else pd.DataFrame()
        
        return fp_inicio, fp_conclusao

    def get_late_and_open_calls(self) -> pd.DataFrame:
        """Obtém chamados atrasados e em aberto"""
        if "Status_Atraso" not in self.df.columns:
            logger.warning("Coluna Status_Atraso não encontrada")
            return pd.DataFrame()
        
        # Filtra apenas chamados atrasados ou em aberto
        late_calls = self.df[
            self.df["Status_Atraso"].isin(["Atrasado", "Concluído com Atraso", "Em Aberto (Sem Previsão)"])
        ].copy()
        
        return late_calls

    def get_accumulated_metrics(self) -> pd.DataFrame:
        """Calcula métricas acumuladas (equivalentes DAX)"""
        if not all(col in self.df.columns for col in ['Data_Criacao', 'Numero_Chamado', 'Data_Conclusao']):
            logger.warning("Colunas necessárias não encontradas para métricas acumuladas")
            return pd.DataFrame()
        
        df_sorted = self.df.sort_values('Data_Criacao')
        
        accumulated = df_sorted.groupby('Data_Criacao').agg(
            Total_Criados=('Numero_Chamado', 'count'),
            Total_Concluidos=('Data_Conclusao', lambda x: x.notna().sum())
        ).cumsum().reset_index()
        
        accumulated['Acumulado_Criados'] = accumulated['Total_Criados']
        accumulated['Acumulado_Concluidos'] = accumulated['Total_Concluidos']
        
        return accumulated

# ======================
# EXPORTAÇÃO EXCEL
# ======================
class ExcelExporter:
    """Exporta análises para Excel com formatação"""

    def __init__(self, output_path: str):
        self.output_path = output_path
        self.writer = None

    def export_analysis(self, processor: STDDataProcessor, analyzer: STDAnalyzer) -> None:
        """Exporta análise completa"""
        logger.info("Exportando análise para Excel...")
        
        try:
            with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                self.writer = writer
                
                # Dados processados
                processor.df_processed.to_excel(writer, sheet_name='Dados_Processados', index=False)
                
                # Estatísticas gerais
                stats = analyzer.calculate_general_stats()
                stats_df = pd.DataFrame(list(stats.items()), columns=['Métrica', 'Valor'])
                stats_df.to_excel(writer, sheet_name='Estatísticas_Gerais', index=False)
                
                # Medidas DAX equivalentes
                dax_measures = self._create_dax_measures_sheet(processor.stats)
                dax_measures.to_excel(writer, sheet_name='Medidas_DAX_Equivalentes', index=False)
                
                # Análises por dimensão
                dimensions = ['Divisão', 'regional', 'Tipo', 'Prioridade', 'Fornecedor', 'UF']
                for dim in dimensions:
                    analysis = analyzer.analyze_by_dimension(dim)
                    if not analysis.empty:
                        sheet_name = f'Por_{dim}'[:31]  # Limite de 31 caracteres
                        analysis.to_excel(writer, sheet_name=sheet_name)
                
                # Evolução mensal
                monthly = analyzer.analyze_monthly_evolution()
                if not monthly.empty:
                    monthly.to_excel(writer, sheet_name='Evolução_Mensal', index=False)
                
                # Top responsáveis
                top_resp = analyzer.get_top_responsibles()
                if not top_resp.empty:
                    top_resp.to_excel(writer, sheet_name='Top_Responsáveis', index=False)
                
                # Análise FP
                fp_inicio, fp_conclusao = analyzer.get_fp_analysis()
                if not fp_inicio.empty:
                    fp_inicio.to_excel(writer, sheet_name='FP_Início', index=False)
                if not fp_conclusao.empty:
                    fp_conclusao.to_excel(writer, sheet_name='FP_Conclusão', index=False)
                
                # Chamados Atrasados e Em Aberto
                late_calls = analyzer.get_late_and_open_calls()
                if not late_calls.empty:
                    late_calls.to_excel(writer, sheet_name='Chamados_Atrasados', index=False)
                
                # Métricas acumuladas
                accumulated = analyzer.get_accumulated_metrics()
                if not accumulated.empty:
                    accumulated.to_excel(writer, sheet_name='Métricas_Acumuladas', index=False)
                
                # Calendário
                if not processor.calendario.empty:
                    processor.calendario.to_excel(writer, sheet_name='Calendario', index=False)
                
                # Aplicar formatação
                self._apply_formatting()
                
            logger.info(f"Análise exportada: {self.output_path}")
            
        except Exception as e:
            logger.error(f"Erro ao exportar Excel: {e}")
            raise

    def _create_dax_measures_sheet(self, stats: Dict[str, float]) -> pd.DataFrame:
        """Cria sheet com medidas DAX equivalentes"""
        dax_measures = [
            {"Medida DAX": "Total Chamados", "Valor": stats["Total Chamados"], "Descrição": "COUNTA('Base WTM'[Numero_Chamado])"},
            {"Medida DAX": "Total Chamados Termino", "Valor": stats["Total Chamados Termino"], "Descrição": "CALCULATE([Total Chamados], USERELATIONSHIP('Base WTM'[Data_Conclusao], Dcalendario[Date]))"},
            {"Medida DAX": "Total Conclusão NP", "Valor": stats["Total Conclusão NP"], "Descrição": "CALCULATE([total chamados termino], 'Base WTM'[Prazo Conclusão Ajustado] = 'NP')"},
            {"Medida DAX": "Total Inicio NP", "Valor": stats["Total Inicio NP"], "Descrição": "CALCULATE([Total Chamados], 'Base WTM'[Prazo Inicio Ajustado] = 'NP')"},
            {"Medida DAX": "SLA Início", "Valor": f"{stats['SLA Início']:.2f}%", "Descrição": "[Total Inicio NP]/[Total Chamados]"},
            {"Medida DAX": "SLA Término", "Valor": f"{stats['SLA Término']:.2f}%", "Descrição": "[Total Conclusão NP]/[total chamados termino]"},
            {"Medida DAX": "Comparação Meta Inicio", "Valor": f"{stats['Comparação Meta Inicio']:.2f} pp", "Descrição": "[SLA Início] - [Meta]"},
            {"Medida DAX": "Comparação Meta Término", "Valor": f"{stats['Comparação Meta Término']:.2f} pp", "Descrição": "[SLA Término] - [Meta]"},
            {"Medida DAX": "Comparação Meta Limpeza Término", "Valor": f"{stats['Comparação Meta Limpeza Término']:.2f} pp", "Descrição": "[SLA Término] - [Meta limpeza]"},
            {"Medida DAX": "Total Estoque", "Valor": stats["Total Estoque"], "Descrição": "CALCULATE([Total Chamados]-[Total Chamados Concluídos])"},
            {"Medida DAX": "Total Fornecedor", "Valor": stats["Total Fornecedor"], "Descrição": "DISTINCTCOUNT('Base WTM'[Fornecedor])"},
            {"Medida DAX": "Total Chamados Concluídos", "Valor": stats["Total Chamados Concluídos"], "Descrição": "CALCULATE([Total Chamados], 'Base WTM'[Data Conclusão Ajustada] <> BLANK())"},
            {"Medida DAX": "Total Chamados FP", "Valor": stats["Total Chamados FP"], "Descrição": "[Total Chamados] - [Total Conclusão NP]"},
            {"Medida DAX": "À VENCER WTM 30 DIAS", "Valor": stats["À VENCER WTM 30 DIAS"], "Descrição": "IF(AND('Base WTM'[DURAÇÃO CHAMADO] < 30, 'Base WTM'[DURAÇÃO CHAMADO] >= 20), 'À VENCER WTM +30 DIAS', 'OUTROS')"},
            {"Medida DAX": "Chamados Atrasados", "Valor": stats["Chamados Atrasados"], "Descrição": "Chamados com prazo vencido ou concluídos com atraso"},
            {"Medida DAX": "Chamados Em Aberto", "Valor": stats["Chamados Em Aberto"], "Descrição": "Chamados sem previsão de conclusão"},
            {"Medida DAX": "Media Dias Atrasos", "Valor": f"{stats['Media Dias Atrasos']:.2f}", "Descrição": "AVERAGE('Base WTM'[Dias atrasos])"},
            {"Medida DAX": "Média Dias Chegada", "Valor": f"{stats['Média Dias Chegada']:.2f}", "Descrição": "CALCULATE(AVERAGE('Base WTM'[Dias Chegada]), 'Base WTM'[Data_Chegada] <> BLANK())"},
            {"Medida DAX": "Média Dias Conclusão", "Valor": f"{stats['Média Dias Conclusão']:.2f}", "Descrição": "CALCULATE(AVERAGE('Base WTM'[Dias Conclusão]), 'Base WTM'[Data_Conclusao] <> BLANK())"},
            {"Medida DAX": "Média Dias Fechamento", "Valor": f"{stats['Média Dias Fechamento']:.2f}", "Descrição": "CALCULATE(AVERAGE('Base WTM'[Dias Fechados]), 'Base WTM'[Data_de_Fechamento] <> BLANK())"},
            {"Medida DAX": "Media Tempo Atendimento", "Valor": f"{stats['Media Tempo Atendimento']:.2f}", "Descrição": "AVERAGE('Base WTM'[Tempo Atendimento])"},
            {"Medida DAX": "Média Valor OS", "Valor": f"{stats['Média Valor OS']:.2f}", "Descrição": "AVERAGE('Base WTM'[Valor_Total])"},
            {"Medida DAX": "Qtd Agencias", "Valor": stats["Qtd Agencias"], "Descrição": "DISTINCTCOUNT('Base WTM'[Uniorg_Comercial])"},
            {"Medida DAX": "Total Valor OS", "Valor": f"{stats['Total Valor OS']:.2f}", "Descrição": "SUM('Base WTM'[Valor_Total])"},
            {"Medida DAX": "Media Tempo Chegada", "Valor": stats["Media Tempo Chegada"], "Descrição": "Formato HHMMSS"},
            {"Medida DAX": "Tempo Chegada", "Valor": stats["Tempo Chegada"], "Descrição": "Formato HHMMSS"}
        ]
        
        return pd.DataFrame(dax_measures)

    def _apply_formatting(self) -> None:
        """Aplica formatação condicional"""
        workbook = self.writer.book
        
        # Formatar planilhas com percentuais
        sheets_to_format = ['Evolução_Mensal', 'Estatísticas_Gerais', 'Medidas_DAX_Equivalentes']
        
        for sheet_name in sheets_to_format:
            if sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                
                # Encontrar colunas de percentual
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value and '%' in str(cell_value):
                        self._format_percentage_column(ws, col)
                    elif cell_value and 'Comparação' in str(cell_value):
                        self._format_comparison_column(ws, col)
        
        # Formatar planilha de chamados atrasados
        if 'Chamados_Atrasados' in workbook.sheetnames:
            ws = workbook['Chamados_Atrasados']
            
            # Encontrar coluna de Status_Atraso
            status_col = None
            dias_atraso_col = None
            
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value == 'Status_Atraso':
                    status_col = col
                elif cell_value == 'Dias_Atraso':
                    dias_atraso_col = col
            
            # Aplicar formatação condicional
            if status_col:
                for row in range(2, ws.max_row + 1):
                    try:
                        status = ws.cell(row=row, column=status_col).value
                        if status == 'Atrasado':
                            for c in range(1, ws.max_column + 1):
                                ws.cell(row=row, column=c).fill = Config.RED_FILL
                        elif status == 'Concluído com Atraso':
                            for c in range(1, ws.max_column + 1):
                                ws.cell(row=row, column=c).fill = Config.ORANGE_FILL
                        elif status == 'Em Aberro (Sem Previsão)':
                            for c in range(1, ws.max_column + 1):
                                ws.cell(row=row, column=c).fill = Config.YELLOW_FILL
                    except Exception:
                        continue

    def _format_percentage_column(self, ws, col_idx: int) -> None:
        """Formata coluna de percentual"""
        for row in range(2, ws.max_row + 1):
            try:
                cell = ws.cell(row=row, column=col_idx)
                value_str = str(cell.value)
                if '%' in value_str:
                    value = float(value_str.replace('%', '').strip())
                else:
                    value = float(cell.value) if cell.value else 0
                
                if 'SLA' in ws.cell(row=1, column=col_idx).value:
                    meta = Config.META_SLA
                else:
                    meta = Config.META_LIMPEZA
                
                if value >= meta:
                    cell.fill = Config.GREEN_FILL
                elif value >= meta - 5:
                    cell.fill = Config.YELLOW_FILL
                else:
                    cell.fill = Config.RED_FILL
                    
            except (ValueError, TypeError):
                continue

    def _format_comparison_column(self, ws, col_idx: int) -> None:
        """Formata coluna de comparação"""
        for row in range(2, ws.max_row + 1):
            try:
                cell = ws.cell(row=row, column=col_idx)
                value_str = str(cell.value)
                if 'pp' in value_str:
                    value = float(value_str.replace('pp', '').strip())
                else:
                    value = float(cell.value) if cell.value else 0
                
                if value >= 0:
                    cell.fill = Config.GREEN_FILL
                else:
                    cell.fill = Config.RED_FILL
                    
            except (ValueError, TypeError):
                continue

# ======================
# MAIN
# ======================
def main():
    """Função principal"""
    try:
        logger.info("Iniciando processamento STD...")
        
        # Processamento
        processor = STDDataProcessor(str(Config.FILE_PATH))
        processor.load_data()
        processor.prepare_data()
        processor.save_processed_data(str(Config.OUTPUT_BASE_TRATADA))
        
        # Análise
        analyzer = STDAnalyzer(processor.df_processed)
        stats = analyzer.calculate_general_stats()
        
        # Exportação
        exporter = ExcelExporter(str(Config.OUTPUT_ANALISE_COMPLETA))
        exporter.export_analysis(processor, analyzer)
        
        # Resultados
        logger.info("Processamento concluído!")
        logger.info(f"Total de chamados: {stats['Total Chamados']:,}")
        logger.info(f"SLA Início: {stats['SLA Início']:.2f}%")
        logger.info(f"SLA Término: {stats['SLA Término']:.2f}%")
        logger.info(f"Estoque atual: {stats['Total Estoque']:,}")
        logger.info(f"Chamados Atrasados: {stats['Chamados Atrasados']:,}")
        logger.info(f"Chamados Em Aberto: {stats['Chamados Em Aberto']:,}")
        logger.info(f"Total Valor OS: R$ {stats['Total Valor OS']:,.2f}")
        
    except Exception as e:
        logger.error(f"Erro no processamento: {e}")
        raise

if __name__ == "__main__":
    main()